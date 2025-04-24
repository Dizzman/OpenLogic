import logging
import sys
import openpyxl
import psycopg2
from config.db_config import DatabaseConfig
from ol_engine.core.ol_engine import OLEngineCore
class LoaderSourceDataToBD():
    def __init__ (self, ol_engeen:OLEngineCore, datafile:str,conffile:str):
        """
        Initialize LoaderSourceData
        :param datafile:
        :param db_cnf:
        """
        self.datafile = datafile
        self.conffile = conffile
        self.ol_engine = ol_engeen
        self.workbook = None
        self.logger = logging.getLogger(self.__class__.__name__)
        #self.logger.info(f"set datafile= {self.datafile}, conffile= {self.conffile} DB name = {self.ol_engine.get_currdb()} Scenario_id= { self.ol_engine.active_scenario_id}")
        self.openconfig_xlsx_file()
        self.opendata_xlsx_file()
        self.logger.info("xls files opened")
       # self.ol_engine.cur.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE';")
        #tables = self.ol_engine.cur.fetchall()

        # Вывод таблиц
        #for table in tables:
        #    print(table[0])
    def opendata_xlsx_file(self):
        """
        Open the Excel file and load the workbook.
        """
        try:
            self.workbook_datafile = openpyxl.load_workbook(self.datafile,data_only=True)
        except FileNotFoundError:
            self.logger.fatal(f"The file at {self.datafile} was not found. ")
            raise FileNotFoundError(f"The file at {self.datafile} was not found.")
        except Exception as e:
            self.logger.fatal(f"An error occurred while opening the file: {e} ")
            raise Exception(f"An error occurred while opening the file: {e}")
    def openconfig_xlsx_file(self):
        """
        Open the Excel file and load the workbook.
        """
        try:
            self.workbook_conffile = openpyxl.load_workbook(self.conffile)
        except FileNotFoundError:
            self.logger.fatal(f"The file at {self.conffile} was not found. ")
            raise FileNotFoundError(f"The file at {self.conffile} was not found.")
        except Exception as e:
            self.logger.fatal(f"An error occurred while opening the file: {e} ")
            raise Exception(f"An error occurred while opening the file: {e}")

    def deleteallsta_tables(self):
        # 1. Получаем список всех таблиц
        self.ol_engine.cur.execute("""
                   SELECT table_name 
                   FROM information_schema.tables 
                   WHERE table_schema = 'public' 
                   AND table_type = 'BASE TABLE'
               """)
        tables = self.ol_engine.cur.fetchall()

        if not tables:
            print("В базе данных нет таблиц для удаления")
            return

        # 2. Формируем и выполняем SQL для удаления всех таблиц
        logging.info(f"Found tables to drop: {len(tables)}")
        for table in tables:
            table_name = table[0]
            try:
                self.ol_engine.cur.execute(f'DROP TABLE IF EXISTS "{table_name}" CASCADE')
                logging.info(f"Table {table_name} successfully dropped")
            except Exception as e:
                logging.error(f"Error dropping table {table_name}: {e}")

        self.ol_engine.conn.commit()


    def get_cell_value(self,sheet, row, col, default=None):
        value = sheet.cell(row=row, column=col).value
        return value if value is not None else default

    def create_Configuration(self):
        logging.info("Create DataTable: Configuration")
        self.ol_engine.cur.execute("""
                CREATE TABLE IF NOT EXISTS T_Configuration (
                    _ScenarioID INT,
                    NumberOfPiles INT,
                    NumberOfVessels INT,
                    NumberOfDiscreteDays INT,
                    NumberOfQCs INT,
                    NumberOfDays INT,
                    ConfirmedDays INT,
                    DVZHDDays INT,
                    FarawayDays INT,
                    PlannedDays INT,
                    NumberOfWeekPeriods INT,
                    NumberOf2WeekPeriods INT,
                    Maxshiftdefault INT);
                """)
        self.ol_engine.conn.commit()
    def create_EO_Tables(self):
        try:
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_E_LocationDefinition.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_E_TimePeriodDefinitions.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_E_LocationDefinitionsMTP.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_C_ConstraintSetDefinitions.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_A_AttributeDefinitions.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_P_PurchaseActivity.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_PtI_MixYield.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_I_InventoryActivity.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_S_SalesActivity.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_ItS_MixDistribution.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_ItI_TransferActivity.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_ItI_FromShipmentDescription.sql")
            self.ol_engine.execute_sql_file("./sql_scripts/EO_Tables/EO_ItI_ToShipmentDescription.sql")

        except Exception as e:
            self.logger.fatal(f"Exit program ")
            exit(1)

    def load_PilesQ(self):
        PilesQ = self.workbook_datafile["quality"]
        for qi in range(1, 1+self.config_data['NumberOfQCs']):
            for pi in range(1, self.config_data.get('NumberOfPiles') + 1):
                PileName = PilesQ.cell(1, 1 + pi).value
                Value = PilesQ.cell(3+qi, 1 + pi).value
                QName = PilesQ.cell(3+qi, 1 ).value
                self.ol_engine.cur.execute("""
                                    SELECT Id FROM T_Piles 
                                    WHERE PileName = %s AND _ScenarioID = %s

                                    """, (PileName, self.ol_engine.active_scenario_id))
                pile_id = self.ol_engine.cur.fetchone()
                print(pile_id,PileName)
                self.ol_engine.cur.execute("""
                                                    SELECT Id FROM T_QltCharacteristics 
                                                    WHERE QualityName = %s AND _ScenarioID = %s
                                                    """, (QName, self.ol_engine.active_scenario_id))
                q_id = self.ol_engine.cur.fetchone()

                self.ol_engine.cur.execute("""
                                            INSERT INTO T_PileQuality
                                            (_ScenarioID,PileId, QualityId, Value) 
                                            VALUES (%s, %s, %s,%s )
                                            """, (
                self.ol_engine.active_scenario_id,  pi,q_id,  Value))
                self.ol_engine.conn.commit()

    def load_Vol(self):
        MinVol = self.workbook_datafile["MinVolume"]
        MaxVol = self.workbook_datafile["MaxVolume"]
        Revenue = self.workbook_datafile["Revenue"]
        for vi in range(1, self.config_data.get('NumberOfVessels') + 1):
            for pi in range(1, self.config_data.get('NumberOfPiles') + 1):
                # Get values from Excel
                PileName = MinVol.cell(1, 3 + pi).value
                VesName = MinVol.cell(3 + vi, 2).value
                mivol = MinVol.cell(3 + vi, 3 + pi).value
                mavol = MaxVol.cell(3 + vi, 3 + pi).value
                rev=Revenue.cell(4 + vi, 8 + pi).value
                # Find vessel_id
                self.ol_engine.cur.execute("""
                    SELECT Id FROM T_Vessels 
                    WHERE VesselName = %s AND _ScenarioID = %s
                    """, (VesName, self.ol_engine.active_scenario_id))
                vessel_id = self.ol_engine.cur.fetchone()

                # Find pile_id (analogous to vessel_id query)
                self.ol_engine.cur.execute("""
                    SELECT Id FROM T_Piles 
                    WHERE PileName = %s AND _ScenarioID = %s
                    """, (PileName, self.ol_engine.active_scenario_id))
                pile_id = self.ol_engine.cur.fetchone()

                # Insert into T_Volumes if both IDs exist
                if vessel_id and pile_id:
                    try:
                        self.ol_engine.cur.execute("""
                            INSERT INTO T_Volumes 
                            (_ScenarioID,VesselId, PileId, MinVolume, MaxVolume,Revenue,VolumePlan) 
                            VALUES (%s, %s, %s, %s , %s, %s, 0 )
                            """, (self.ol_engine.active_scenario_id,vessel_id[0], pile_id[0], mivol, mavol,rev))
                        self.ol_engine.conn.commit()
                    except Exception as e:
                        self.ol_engine.conn.rollback()
                        print(f"Error inserting volume data: {str(e)}")

    def load_Q(self):
        sheet_Q = self.workbook_datafile["Q"]
        self.logger.info("Loading Q data into T_QltCharacteristics")

        try:
            for i in range(1, 1+self.config_data['NumberOfQCs']):
                QualityCode = f"QLT{i:03d}"
                QualityName = sheet_Q.cell(3, 4 + i).value


                # Insert record into database
                try:
                    self.ol_engine.cur.execute("""
                        INSERT INTO T_QltCharacteristics (_ScenarioID, QualityCode, QualityName)
                        VALUES (%s, %s, %s)
                        """, (self.ol_engine.active_scenario_id, QualityCode, QualityName))

                    # Commit after each successful insert
                    self.ol_engine.conn.commit()


                except Exception as e:
                    self.ol_engine.conn.rollback()
                    print(f"Error inserting record {i} ({QualityCode}): {str(e)}")
                    # Continue with next record even if one fails
                    continue

        except Exception as e:
            print(f"Critical error in load_Q: {str(e)}")
            raise  # Re-raise the exception if it's a major error

        for qi in range(1, 1+self.config_data['NumberOfQCs']):  # First 99 quality characteristics
            for vi in range(1, self.config_data['NumberOfVessels'] + 1):
                try:
                    # Get vessel and quality info from Excel
                    vesname = sheet_Q.cell(3 + vi, 1).value
                    qname = sheet_Q.cell(3, 4 + qi).value
                    qmin = sheet_Q.cell(3 + vi, 4 + qi).value
                    qmax = sheet_Q.cell(3 + vi, 107 + qi).value  # Assuming max values start at column 107

                    # Skip if no values
                    if qmin is None and qmax is None:
                        continue

                    # Get VesselId
                    self.ol_engine.cur.execute("""
                        SELECT Id FROM T_Vessels 
                        WHERE VesselName = %s AND _ScenarioID  = %s
                        """, (vesname, self.ol_engine.active_scenario_id))
                    vessel_id = self.ol_engine.cur.fetchone()
                    if not vessel_id:
                        print(f"Vessel not found: {vesname}")
                        continue
                    vessel_id = vessel_id[0]

                    # Get QualityCharacteristicId
                    self.ol_engine.cur.execute("""
                        SELECT Id FROM T_QltCharacteristics 
                        WHERE QualityName = %s AND _ScenarioID = %s
                        """, (qname, self.ol_engine.active_scenario_id))
                    qlt_id = self.ol_engine.cur.fetchone()
                    if not qlt_id:
                        print(f"Quality characteristic not found: {qname}")
                        continue
                    qlt_id = qlt_id[0]

                    # Insert restriction
                    self.ol_engine.cur.execute("""
                        INSERT INTO T_QualityRestrictions 
                        (VesselId,_ScenarioID, QltCharacteristicId, ValueMin, ValueMax)
                        VALUES (%s, %s,%s, %s, %s)
                        """, (vessel_id,self.ol_engine.active_scenario_id, qlt_id, qmin, qmax))
                    self.ol_engine.conn.commit()

                except Exception as e:
                    self.ol_engine.conn.rollback()
                    print(f"Error processing vessel {vi} quality {qi}: {str(e)}")
                    continue


        print(f"Successfully  Quality")





    def load_cargo_data(self):
        sheet_cargo = self.workbook_datafile["Cargo"]

        # Проверка существующих записей
        self.ol_engine.cur.execute("""
            SELECT COUNT(*) FROM T_Piles 
            WHERE _ScenarioID = %s
        """, (self.ol_engine.active_scenario_id,))
        count = self.ol_engine.cur.fetchone()[0]

        if count > 0:
            self.logger.info("Found %d records with _ScenarioID = %d. Deleting...",
                             count, self.ol_engine.active_scenario_id)
            self.ol_engine.cur.execute("""
                DELETE FROM T_Piles 
                WHERE _ScenarioID = %s
            """, (self.ol_engine.active_scenario_id,))
            self.logger.info("Deleted %d records with _ScenarioID = %d",
                             count, self.ol_engine.active_scenario_id)

        self.logger.info("Loading Cargo data into T_Piles")

        # Количество штабелей (предположим, что оно задано в конфиге)
        number_of_piles = self.config_data.get('NumberOfPiles', 30)

        # Чтение данных из Excel
        for pile_number in range(1, number_of_piles + 1):
            pile_code = 'Pile'+str(self.get_cell_value(sheet_cargo, 2, 3 + pile_number, ""))

            pile_name = str(self.get_cell_value(sheet_cargo, 1, 3 + pile_number, ""))

            min_stock = round(float(self.get_cell_value(sheet_cargo, 3, 3 + pile_number, 0)), 2)
            stock_yard = round(float(self.get_cell_value(sheet_cargo, 4, 3 + pile_number, 0)), 2)
            on_board = round(float(self.get_cell_value(sheet_cargo, 5, 3 + pile_number, 0)), 2)

            confirmed_volume = round(float(self.get_cell_value(sheet_cargo, 6, 3 + pile_number, 0)), 2)
            confirmed_days = round(float(self.get_cell_value(sheet_cargo, 7, 3 + pile_number, 0)), 2)

            dvzhd_volume = round(float(self.get_cell_value(sheet_cargo, 8, 3 + pile_number, 0)), 2)
            dvzhd_days = round(float(self.get_cell_value(sheet_cargo, 9, 3 + pile_number, 0)), 2)

            faraway_volume = round(float(self.get_cell_value(sheet_cargo, 10, 3 + pile_number, 0)), 2)
            faraway_days = round(float(self.get_cell_value(sheet_cargo, 11, 3 + pile_number, 0)), 2)

            planned_volume = round(float(self.get_cell_value(sheet_cargo, 12, 3 + pile_number, 0)), 2)
            planned_days = round(float(self.get_cell_value(sheet_cargo, 13, 3 + pile_number, 0)), 2)

            longrun_volume = round(float(self.get_cell_value(sheet_cargo, 14, 3 + pile_number, 0)), 2)
            longrun_days = round(float(self.get_cell_value(sheet_cargo, 15, 3 + pile_number, 0)), 2)

            # Вставка данных в таблицу
            insert_query = """
            INSERT INTO T_Piles (
                _ScenarioID, PileCode, PileName, 
                MinStock, StockYard, OnBoard,
                ConfirmedVolume, ConfirmedDays,
                DVZHDVolume, DVZHDDays,
                FarawayVolume, FarawayDays,
                PlannedVolume, PlannedDays,
                LongRunVolume, LongRunDays
            ) VALUES (
                %s, %s, %s, 
                %s, %s, %s,
                %s, %s,
                %s, %s,
                %s, %s,
                %s, %s,
                %s, %s
            )
            """

            self.ol_engine.cur.execute(insert_query, (
                self.ol_engine.active_scenario_id,
                pile_code,
                pile_name,
                min_stock,
                stock_yard,
                on_board,
                confirmed_volume,
                confirmed_days,
                dvzhd_volume,
                dvzhd_days,
                faraway_volume,
                faraway_days,
                planned_volume,
                planned_days,
                longrun_volume,
                longrun_days
            ))

        self.ol_engine.conn.commit()
        self.logger.info("Successfully loaded %d piles into T_Piles", number_of_piles)

    def load_vessels_from_xls_to_db(self):
        sheet_vessels = self.workbook_datafile["Vessels"]

        # Check for records with _ScenarioID = 1
        self.ol_engine.cur.execute("""
                       SELECT COUNT(*) FROM T_Vessels
                       WHERE _ScenarioID= %s
                       """, (self.ol_engine.active_scenario_id,))
        count = self.ol_engine.cur.fetchone()[0]  # Get the count of records

        if count > 0:
            self.logger.info("Found %d records with _ScenarioID = %d. Proceeding to delete.", count,
                             self.ol_engine.active_scenario_id)
            # Delete records where _ScenarioID = 1
            self.ol_engine.cur.execute("""
                           DELETE FROM T_Vessels 
                           WHERE _ScenarioID = %s
                           """, (self.ol_engine.active_scenario_id,))
            self.logger.info(f"Records with _ScenarioID = %d deleted.", self.ol_engine.active_scenario_id)
        else:
            self.logger.info("No records found with _ScenarioID = %d. Delete not required.",
                             self.ol_engine.active_scenario_id)
        self.logger.info("Load Vessels data")

        # Цикл по строкам таблицы
        for row_number in range(1,  self.config_data['NumberOfVessels']+1):  # Начинаем с 2-й строки, т.к. первая - заголовок
            fix_value = int(self.get_cell_value(sheet_vessels, row_number+1, 12, 0))
            days_before_arrival = int(self.get_cell_value(sheet_vessels, row_number+1, 1, 0)) if fix_value == 0 else None
            vessel_name = str(self.get_cell_value(sheet_vessels, row_number+1, 3, ''))
            customer = str(self.get_cell_value(sheet_vessels, row_number+1, 4, ''))
            grade = str(self.get_cell_value(sheet_vessels, row_number+1, 5, ''))
            demurrage = round(float(self.get_cell_value(sheet_vessels, row_number+1, 6, 0)), 4) if fix_value == 0 else None
            max_load_speed_contract =round(float(self.get_cell_value(sheet_vessels, row_number+1, 7, 0)),4) if fix_value == 0 else None
            max_load_speed_real = round(float(self.get_cell_value(sheet_vessels, row_number+1, 8, 0)), 4) if fix_value == 0 else None
            volume_min = round(float(self.get_cell_value(sheet_vessels, row_number+1, 9, 0)), 4) if fix_value == 0 else None
            volume_max = round(float(self.get_cell_value(sheet_vessels, row_number+1, 10, 0)), 4) if fix_value == 0 else None
            contract = str(self.get_cell_value(sheet_vessels, row_number+1, 11, ''))
            basis = str(self.get_cell_value(sheet_vessels, row_number+1, 2, ''))
            max_shift_days = int(self.get_cell_value(sheet_vessels, row_number+1, 13, 0))

            # Вставка данных в таблицу PostgreSQL
            insert_query = """
            INSERT INTO T_Vessels  (
                _ScenarioID, VesselCode, DaysBeforeArrival, VesselName,  Customer, Grade, Demurrage, MaxLoadSpeedContract, MaxLoadSpeedReal, VolumeMin, VolumeMax, Contract,FIX ,Days,Basis, Maxshift
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s
            );
            """
            self.ol_engine.cur.execute(insert_query, (
                self.ol_engine.active_scenario_id,  # ConfigId - если отсутствует, вставляется NULL
                f"Ves{row_number:02}",
                days_before_arrival,
                vessel_name,
                customer,
                grade,
                demurrage,
                max_load_speed_contract,
                max_load_speed_real,
                volume_min,
                volume_max,
                contract,
                fix_value,
                0.0,
                basis,
                max_shift_days
            ))

        self.ol_engine.conn.commit()
    def load_configure_from_xls_to_db(self):
        """
        Save data from into the database.

        :param sheet_name: Name of the sheet_config to save data from
        :param table_name: Name of the database table to insert data into
        """

        sheet_config = self.workbook_conffile["Config"]
        sheet_vessels = self.workbook_datafile["Vessels"]
        sheet_cargo = self.workbook_datafile["Cargo"]
        self.config_data = {
             'NumberOfPiles': sheet_config.cell(row=2, column=3).value,
             'NumberOfVessels': sheet_config.cell(row=3, column=3).value,
             'NumberOfDiscreteDays': sheet_config.cell(row=4, column=3).value,
             'NumberOfQCs': sheet_config.cell(row=5, column=3).value,
             'NumberOfDays': sheet_config.cell(row=6, column=3).value,
             'ConfirmedDays': sheet_cargo.cell(row=7, column=1).value,
             'DVZHDDays': sheet_cargo.cell(row=9, column=1).value,
             'FarawayDays': sheet_cargo.cell(row=11, column=1).value,
             'PlannedDays': sheet_cargo.cell(row=13, column=1).value,
             'NumberOfWeekPeriods': sheet_config.cell(row=7, column=3).value,
             'NumberOf2WeekPeriods': sheet_config.cell(row=8, column=3).value,
             'Maxshiftdefault': sheet_config.cell(row=10, column=3).value # Запись можно брать из Vessels
        }
        self.logger.info("Data extracted: %s", self.config_data)
        # Create table if it does not exist

        self.logger.info("Configuration table checked/created.")

        # Check for records with _ScenarioID = 1
        self.ol_engine.cur.execute("""
                SELECT COUNT(*) FROM T_Configuration 
                WHERE _ScenarioID = %s
                """, (self.ol_engine.active_scenario_id ,))
        count = self.ol_engine.cur.fetchone()[0]  # Get the count of records

        if count > 0:
            self.logger.info("Found %d records with _ScenarioID = %d. Proceeding to delete.", count,
                             self.ol_engine.active_scenario_id)
            # Delete records where _ScenarioID = 1
            self.ol_engine.cur.execute("""
                    DELETE FROM T_Configuration 
                    WHERE _ScenarioID = %s
                    """, (self.ol_engine.active_scenario_id,))
            self.logger.info(f"Records with _ScenarioID = %d deleted.", self.ol_engine.active_scenario_id)
        else:
            self.logger.info("No records found with _ScenarioID = %d. Deletion not required.",
                             self.ol_engine.active_scenario_id)

        # Insert data with explicit enumeration
        self.ol_engine.cur.execute("""
                INSERT INTO T_Configuration (_ScenarioID,NumberOfPiles, NumberOfVessels, NumberOfDiscreteDays, NumberOfQCs, 
                                           NumberOfDays, ConfirmedDays, DVZHDDays, FarawayDays, PlannedDays,NumberOfWeekPeriods,NumberOf2WeekPeriods,Maxshiftdefault) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
            self.ol_engine.active_scenario_id,
            self.config_data['NumberOfPiles'],
            self.config_data['NumberOfVessels'],
            self.config_data['NumberOfDiscreteDays'],
            self.config_data['NumberOfQCs'],
            self.config_data['NumberOfDays'],
            self.config_data['ConfirmedDays'],
            self.config_data['DVZHDDays'],
            self.config_data['FarawayDays'],
            self.config_data['PlannedDays'],
            self.config_data['NumberOfWeekPeriods'],
            self.config_data['NumberOf2WeekPeriods'],
            self.config_data['Maxshiftdefault']

        ))

        self.logger.info("Data successfully inserted into the Configuration table.")
        self.ol_engine.conn.commit()
        pass
    # def saveconffile_to_database(self):
    #     """
    #     Save data from into the database.
    #
    #     :param sheet_name: Name of the sheet to save data from
    #     :param table_name: Name of the database table to insert data into
    #     """
    #     self.openconfig_xlsx_file()
    #     sheet = self.workbook_conffile["Config"]
    #     pass
    def getall_proc(self):
        self.ol_engine.cur.execute(
                f"SELECT proname AS procedure_name "
                f"FROM pg_proc "
                f"WHERE pronamespace = (SELECT oid FROM pg_namespace WHERE nspname = 'public') "
                f"ORDER BY procedure_name;"
            )
        proclist = self.ol_engine.cur.fetchall()
        return proclist

    def delete_all_procedures(self):
        """Delete all stored procedures and functions in the public schema"""
        try:
            # Get list of all functions/procedures with their full signatures
            self.ol_engine.cur.execute("""
                SELECT 
                    CASE WHEN p.prokind = 'p' THEN 'PROCEDURE' ELSE 'FUNCTION' END as obj_type,
                    'public.' || p.proname || '(' || 
                    COALESCE(pg_get_function_identity_arguments(p.oid), '') || ')' as obj_signature
                FROM pg_proc p
                JOIN pg_namespace n ON p.pronamespace = n.oid
                WHERE n.nspname = 'public'
                ORDER BY p.proname;
            """)

            objects = self.ol_engine.cur.fetchall()
            self.logger.info(f"Found {len(objects)} functions/procedures in public schema")

            # If nothing found, use default list (consider removing this in production)
            if not objects:
                default_objects = [
                    ('FUNCTION', 'public.eosms_run_all_sp(int4)'),
                    ('FUNCTION', 'public.eotemp_activevessel_sp(int4)'),
                    ('FUNCTION', 'public.eo_e_timeperioddefinitions_sp(int8)')
                ]
                objects = default_objects
                self.logger.warning("No procedures found, using default list")

            # Delete each object
            success_count = 0
            for obj_type, signature in objects:
                try:
                    drop_sql = f"DROP {obj_type} IF EXISTS {signature} CASCADE"
                    self.logger.debug(f"Executing: {drop_sql}")
                    self.ol_engine.cur.execute(drop_sql)
                    success_count += 1
                    self.logger.info(f"Successfully dropped: {signature}")
                except Exception as proc_error:
                    self.logger.error(f"Failed to drop {signature}: {str(proc_error)}")
                    continue

            self.ol_engine.conn.commit()
            self.logger.info(
                f"Procedure deletion completed. Successfully dropped {success_count} of {len(objects)} objects")

        except Exception as main_error:
            self.logger.error(f"Error in delete_all_procedures: {str(main_error)}")
            self.ol_engine.conn.rollback()
            raise



    def add_sql_procedure_from_file(self, file_path):
        self.ol_engine.add_sql_procedure_from_file(file_path)

    def call_sql_procedure_for_active_scenario(self, procedure_name, *args):
        """
        Вызывает SQL-процедуру для активного сценария.

        :param procedure_name: Имя процедуры.
        :param args: Аргументы процедуры (если есть).
        """
        try:
            # Формируем запрос для вызова процедуры
            query = f"CALL {procedure_name}(%s);"

            # Выполняем запрос с активным scenario_id
            self.ol_engine.cur.execute(query, (self.ol_engine.active_scenario_id,))

            # Фиксируем изменения
            self.ol_engine.conn.commit()

            # Логируем успешный вызов
            self.logger.info(f"Процедура {procedure_name} успешно вызвана.")
        except Exception as e:
            # Логируем ошибку и откатываем транзакцию
            self.logger.error(f"Ошибка вызова процедуры {procedure_name}: {e}")
            self.ol_engine.conn.rollback()

    def execute_sql_file(self,pathsql):
            self.ol_engine.execute_sql_file(pathsql)