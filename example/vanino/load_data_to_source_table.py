import logging
import sys
import openpyxl
import psycopg2
from config.db_config import DatabaseConfig
from ol_engine.core.ol_engine import OLEngineCore
class LoaderSourceDataToBD():
    def __init__ (self, ol_engeen:OLEngineCore, datafile:str,conffile:str):
        """
        Initialize LoaderSourceData
        :param datafile:
        :param db_cnf:
        """
        self.datafile = datafile
        self.conffile = conffile
        self.ol_engine = ol_engeen
        self.workbook = None
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.info(f"set datafile= {self.datafile}, conffile= {self.conffile} DB name = {self.ol_engine.get_currdb()} Scenario_id= { self.ol_engine.active_scenario_id}")
        self.openconfig_xlsx_file()
        self.opendata_xlsx_file()
        self.logger.info("xls files opened")
       # self.ol_engine.cur.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE';")
        #tables = self.ol_engine.cur.fetchall()

        # Вывод таблиц
        #for table in tables:
        #    print(table[0])
    def opendata_xlsx_file(self):
        """
        Open the Excel file and load the workbook.
        """
        try:
            self.workbook_datafile = openpyxl.load_workbook(self.datafile,data_only=True)
        except FileNotFoundError:
            self.logger.fatal(f"The file at {self.datafile} was not found. ")
            raise FileNotFoundError(f"The file at {self.datafile} was not found.")
        except Exception as e:
            self.logger.fatal(f"An error occurred while opening the file: {e} ")
            raise Exception(f"An error occurred while opening the file: {e}")
    def openconfig_xlsx_file(self):
        """
        Open the Excel file and load the workbook.
        """
        try:
            self.workbook_conffile = openpyxl.load_workbook(self.conffile)
        except FileNotFoundError:
            self.logger.fatal(f"The file at {self.conffile} was not found. ")
            raise FileNotFoundError(f"The file at {self.conffile} was not found.")
        except Exception as e:
            self.logger.fatal(f"An error occurred while opening the file: {e} ")
            raise Exception(f"An error occurred while opening the file: {e}")

    def deleteallsta_tables(self):
        tables = ['T_Configuration','T_Vessels','T_EOtemp_ActiveVessel']
        for i in tables:
            self.ol_engine.cur.execute(f"DROP TABLE IF EXISTS {i}")
        self.ol_engine.conn.commit()


    def get_cell_value(self,sheet, row, col, default=None):
        value = sheet.cell(row=row, column=col).value
        return value if value is not None else default

    def createt_T_EOtemp_ActiveVessel(self):
        self.ol_engine.cur.execute("""CREATE TABLE IF NOT EXISTS T_EOtemp_ActiveVessel  
        (
    _ScenarioID          INT,
    vesselid          INT,
    vesselcode        VARCHAR,
    eo_vessel         VARCHAR,
    dayofstart        INT,
    daystoload        INT,
    discretedaysves   INT,
    demur_ks          NUMERIC,
    maxspeed_kt       NUMERIC,
    volmin_kt         NUMERIC,
    volmax_kt         NUMERIC,
    maxshiftdays      INT);""")
        self.ol_engine.conn.commit()
    def load_vessels_from_xls_to_db(self):
        sheet_vessels = self.workbook_datafile["Vessels"]

        self.ol_engine.cur.execute("""
                             CREATE TABLE IF NOT EXISTS T_Vessels (                           
                                  _ScenarioID INT,
                                  Id   serial primary key,
                                  VesselCode VARCHAR,
                                  DaysBeforeArrival INT,
                                  VesselName VARCHAR,
                                  Customer VARCHAR,
                                  Grade VARCHAR,
                                  Demurrage NUMERIC,
                                  MaxLoadSpeedContract NUMERIC,
                                  MaxLoadSpeedReal NUMERIC,
                                  VolumeMin NUMERIC,
                                  VolumeMax NUMERIC,
                                  Contract VARCHAR,
                                  FIX INT,
                                  Days NUMERIC,
                                  Basis VARCHAR,
                                  Maxshift INT);
                             """)
        # Check for records with _ScenarioID = 1
        self.ol_engine.cur.execute("""
                       SELECT COUNT(*) FROM T_Vessels 
                       WHERE _ScenarioID= %s
                       """, (self.ol_engine.active_scenario_id,))
        count = self.ol_engine.cur.fetchone()[0]  # Get the count of records

        if count > 0:
            self.logger.info("Found %d records with _ScenarioID = %d. Proceeding to delete.", count,
                             self.ol_engine.active_scenario_id)
            # Delete records where _ScenarioID = 1
            self.ol_engine.cur.execute("""
                           DELETE FROM T_Vessels 
                           WHERE _ScenarioID = %s
                           """, (self.ol_engine.active_scenario_id,))
            self.logger.info(f"Records with _ScenarioID = %d deleted.", self.ol_engine.active_scenario_id)
        else:
            self.logger.info("No records found with _ScenarioID = %d. Deletion not required.",
                             self.ol_engine.active_scenario_id)
        self.logger.info("Load Vessels data")

        # Цикл по строкам таблицы
        for row_number in range(1,  self.config_data['NumberOfVessels']+1):  # Начинаем с 2-й строки, т.к. первая - заголовок
            fix_value = int(self.get_cell_value(sheet_vessels, row_number+1, 12, 0))
            days_before_arrival = int(self.get_cell_value(sheet_vessels, row_number+1, 1, 0)) if fix_value == 0 else None
            vessel_name = str(self.get_cell_value(sheet_vessels, row_number+1, 3, ''))
            customer = str(self.get_cell_value(sheet_vessels, row_number+1, 4, ''))
            grade = str(self.get_cell_value(sheet_vessels, row_number+1, 5, ''))
            demurrage = round(float(self.get_cell_value(sheet_vessels, row_number+1, 6, 0)), 4) if fix_value == 0 else None
            max_load_speed_contract =round(float(self.get_cell_value(sheet_vessels, row_number+1, 7, 0)),4) if fix_value == 0 else None
            max_load_speed_real = round(float(self.get_cell_value(sheet_vessels, row_number+1, 8, 0)), 4) if fix_value == 0 else None
            volume_min = round(float(self.get_cell_value(sheet_vessels, row_number+1, 9, 0)), 4) if fix_value == 0 else None
            volume_max = round(float(self.get_cell_value(sheet_vessels, row_number+1, 10, 0)), 4) if fix_value == 0 else None
            contract = str(self.get_cell_value(sheet_vessels, row_number+1, 11, ''))
            basis = str(self.get_cell_value(sheet_vessels, row_number+1, 2, ''))
            max_shift_days = int(self.get_cell_value(sheet_vessels, row_number+1, 13, 0))

            # Вставка данных в таблицу PostgreSQL
            insert_query = """
            INSERT INTO T_Vessels  (
                _ScenarioID, VesselCode, DaysBeforeArrival, VesselName,  Customer, Grade, Demurrage, MaxLoadSpeedContract, MaxLoadSpeedReal, VolumeMin, VolumeMax, Contract,FIX ,Days,Basis, Maxshift
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s
            );
            """
            self.ol_engine.cur.execute(insert_query, (
                self.ol_engine.active_scenario_id,  # ConfigId - если отсутствует, вставляется NULL
                f"Ves{row_number:02}",
                days_before_arrival,
                vessel_name,
                customer,
                grade,
                demurrage,
                max_load_speed_contract,
                max_load_speed_real,
                volume_min,
                volume_max,
                contract,
                fix_value,
                0.0,
                basis,
                max_shift_days
            ))

        self.ol_engine.conn.commit()
    def load_configure_from_xls_to_db(self):
        """
        Save data from into the database.

        :param sheet_name: Name of the sheet_config to save data from
        :param table_name: Name of the database table to insert data into
        """

        sheet_config = self.workbook_conffile["Config"]
        sheet_vessels = self.workbook_datafile["Vessels"]
        sheet_cargo = self.workbook_datafile["Cargo"]
        self.config_data = {
             'NumberOfPiles': sheet_config.cell(row=2, column=3).value,
             'NumberOfVessels': sheet_config.cell(row=3, column=3).value,
             'NumberOfDiscreteDays': sheet_config.cell(row=4, column=3).value,
             'NumberOfQCs': sheet_config.cell(row=5, column=3).value,
             'NumberOfDays': sheet_config.cell(row=6, column=3).value,
             'ConfirmedDays': sheet_cargo.cell(row=7, column=1).value,
             'DVZHDDays': sheet_cargo.cell(row=9, column=1).value,
             'FarawayDays': sheet_cargo.cell(row=11, column=1).value,
             'PlannedDays': sheet_cargo.cell(row=13, column=1).value,
             'NumberOfWeekPeriods': sheet_config.cell(row=7, column=3).value,
             'NumberOf2WeekPeriods': sheet_config.cell(row=8, column=3).value,
             'Maxshiftdefault': sheet_config.cell(row=10, column=3).value # Запись можно брать из Vessels
        }
        self.logger.info("Data extracted: %s", self.config_data)
        # Create table if it does not exist
        self.ol_engine.cur.execute("""
                CREATE TABLE IF NOT EXISTS T_Configuration (
                    _ScenarioID INT,
                    NumberOfPiles INT,
                    NumberOfVessels INT,
                    NumberOfDiscreteDays INT,
                    NumberOfQCs INT,
                    NumberOfDays INT,
                    ConfirmedDays INT,
                    DVZHDDays INT,
                    FarawayDays INT,
                    PlannedDays INT,
                    NumberOfWeekPeriods INT,
                    NumberOf2WeekPeriods INT,
                    Maxshiftdefault INT);
                """)
        self.logger.info("Configuration table checked/created.")

        # Check for records with _ScenarioID = 1
        self.ol_engine.cur.execute("""
                SELECT COUNT(*) FROM T_Configuration 
                WHERE _ScenarioID = %s
                """, (self.ol_engine.active_scenario_id ,))
        count = self.ol_engine.cur.fetchone()[0]  # Get the count of records

        if count > 0:
            self.logger.info("Found %d records with _ScenarioID = %d. Proceeding to delete.", count,
                             self.ol_engine.active_scenario_id)
            # Delete records where _ScenarioID = 1
            self.ol_engine.cur.execute("""
                    DELETE FROM T_Configuration 
                    WHERE _ScenarioID = %s
                    """, (self.ol_engine.active_scenario_id,))
            self.logger.info(f"Records with _ScenarioID = %d deleted.", self.ol_engine.active_scenario_id)
        else:
            self.logger.info("No records found with _ScenarioID = %d. Deletion not required.",
                             self.ol_engine.active_scenario_id)

        # Insert data with explicit enumeration
        self.ol_engine.cur.execute("""
                INSERT INTO T_Configuration (_ScenarioID,NumberOfPiles, NumberOfVessels, NumberOfDiscreteDays, NumberOfQCs, 
                                           NumberOfDays, ConfirmedDays, DVZHDDays, FarawayDays, PlannedDays,NumberOfWeekPeriods,NumberOf2WeekPeriods,Maxshiftdefault) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
            self.ol_engine.active_scenario_id,
            self.config_data['NumberOfPiles'],
            self.config_data['NumberOfVessels'],
            self.config_data['NumberOfDiscreteDays'],
            self.config_data['NumberOfQCs'],
            self.config_data['NumberOfDays'],
            self.config_data['ConfirmedDays'],
            self.config_data['DVZHDDays'],
            self.config_data['FarawayDays'],
            self.config_data['PlannedDays'],
            self.config_data['NumberOfWeekPeriods'],
            self.config_data['NumberOf2WeekPeriods'],
            self.config_data['Maxshiftdefault']

        ))

        self.logger.info("Data successfully inserted into the Configuration table.")
        self.ol_engine.conn.commit()
        pass
    # def saveconffile_to_database(self):
    #     """
    #     Save data from into the database.
    #
    #     :param sheet_name: Name of the sheet to save data from
    #     :param table_name: Name of the database table to insert data into
    #     """
    #     self.openconfig_xlsx_file()
    #     sheet = self.workbook_conffile["Config"]
    #     pass
    def delete_all_procedures(self):
        proc_list=['public.eosms_run_all_sp(int4)','public.eotemp_activevessel_sp(int4)']
        for i in proc_list:
            sql = f"DROP FUNCTION IF EXISTS {i} "
            self.ol_engine.cur.execute(sql)
        self.ol_engine.conn.commit()

    def execute_sql_file(self, file_path):
        try:
            with open(file_path, 'r',encoding="utf-8") as file:
                sql_script = file.read()
                self.ol_engine.cur.execute(sql_script)
                self.ol_engine.cur.execute("SELECT public.eosms_run_all_sp(%s);", (self.ol_engine.active_scenario_id,))
                self.ol_engine.conn.commit()   # Commit changes
                self.logger.info(f"SQL procedure {file_path} added")
        except Exception as e:
            self.logger.info(f"Error executed SQL файла: {e}")
            self.ol_engine.conn.rollback()  # Rollback changes in case of error

    def add_sql_procedure_from_file(self, file_path):
        try:
            with open(file_path, 'r',encoding="utf-8") as file:
                sql_script = file.read()
                self.ol_engine.cur.execute(sql_script)
                self.ol_engine.conn.commit()   # Commit changes
                self.logger.info(f"SQL procedure {file_path} added")
        except Exception as e:
            self.logger.info(f"Error executed SQL файла: {e}")
            self.ol_engine.conn.rollback()  # Rollback changes in case of error
