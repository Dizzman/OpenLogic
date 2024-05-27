import logging
import traceback
from openpyxl import load_workbook
from Obj_Types import OBJECT_TYPE

class OLBaseObject(object):
    # Constructor
    def __init__(self,BaseModel,ObjName):
        self.BaseModel = BaseModel
        self.ObjName = ObjName
        self.ObjType = OBJECT_TYPE.NONETYPE
    def __AddRow(self,ObjName,MapNameValues):
        strnames = ''
        strvalues = ''

        for i in MapNameValues.keys():
            strnames=strnames +','+i
            if type(MapNameValues[i])==str:
                strvalues = strvalues + ',' + '"'+str(MapNameValues[i])+'"'
            else:
                strvalues = strvalues + ',' + str(MapNameValues[i])

        if self.ObjType==OBJECT_TYPE.MixPurch2Inv or self.ObjType==OBJECT_TYPE.MixInv2Sale:
            sql = f'''INSERT INTO {self.m_nameTable}(ObjectName,FromObjectName,ToObjectName,_sessionID{strnames})
                                 VALUES("{ObjName}","{self.FromObjectName}","{self.ToObjectName}",{self.BaseModel.m_session_id}{strvalues}) '''
        else:
            sql = f'''INSERT INTO {self.m_nameTable}(ObjectName,_sessionID{strnames})
                     VALUES("{ObjName}",{self.BaseModel.m_session_id}{strvalues}) '''
        self.BaseModel.cursor.execute(sql)
        pass

    def AddObject_Rows(self, xlsxfile):
        wb = load_workbook(xlsxfile)
        sheet = wb.get_sheet_by_name('Sheet1')
        MapNameValues= {}
        if self.ObjType==OBJECT_TYPE.NONETYPE:
            logging.error(f"Unknkown ObjeType{__name__} {__file__} ")
            raise Exception(f"Unknkown ObjeType{__name__} {__file__} ")
        for ir in range(sheet.max_row-1):
            MapNameValues.clear()
            for ic in range(sheet.max_column):
                 namecolRL=sheet.cell(row=1, column=ic+1).value
                 namecolOL=self.m_mapper.getRL2OL(namecolRL)
                 if namecolOL is None: # This columnnane not used in OL Engeen
                    continue
                 value = sheet.cell(row=ir+2, column=ic + 1).value
                 MapNameValues[namecolOL]=value
            self.__AddRow(self.ObjName,MapNameValues)
        self.BaseModel.connection.commit()