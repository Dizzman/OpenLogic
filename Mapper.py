from openpyxl import load_workbook
import os
class Mapper:
    def __init__(self,listOLcolumns):
        self.listOLcolumns = listOLcolumns
    def getOL2RL(self,OL):
        if OL in self.MapOL2RL:
            return self.MapOL2RL[OL]
        return None

    def getRL2OL(self, RL):
        if RL in self.MapRL2OL:
            return self.MapRL2OL[RL]
        return None

    def openxlsmap(self,filename):
        self.MapOL2RL = {}
        self.MapRL2OL = {}
        self.bpath = os.path.dirname(os.path.realpath(__file__))
        wb = load_workbook(self.bpath+"\\mapping\\"+filename)
        sheet = wb.get_sheet_by_name('Map')

        for i in range(sheet.max_row):
            RLName = sheet.cell(row=i + 1, column=1).value
            OLName = sheet.cell(row=i + 1, column=2).value
            if not OLName==None and  not OLName in self.listOLcolumns:
                raise NameError(f'not corrent OL column name {OLName}  in file {filename}')
            if not OLName == None:
                self.MapOL2RL[OLName] = RLName
                self.MapRL2OL[RLName] = OLName
